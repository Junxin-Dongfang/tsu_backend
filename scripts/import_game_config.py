#!/usr/bin/env python3
"""
游戏配置表导入工具
从 Excel 文件导入游戏配置到数据库

使用方法:
    python3 scripts/import_game_config.py --file configs/game/游戏配置表_v2.0.0.xlsx
"""

import openpyxl
import psycopg2
import json
import argparse
import sys
from datetime import datetime
from typing import Dict, List, Any, Optional

# 数据库配置（默认容器内连接）
DB_CONFIG = {
    'host': '127.0.0.1',  # 容器内使用127.0.0.1可以trust认证
    'port': 5432,
    'database': 'tsu_db',
    'user': 'tsu_user',
    'password': ''  # 容器内127.0.0.1无密码认证
}

# Sheet 到数据库表的映射
SHEET_MAPPING = {
    '角色数据类型': 'game_config.hero_attribute_type',  # 注意：单数
    '伤害类型': 'game_config.damage_types',
    '动作类别': 'game_config.action_categories',
    '特征配置': 'game_config.tags',
    '元效果类型定义': 'game_config.effect_type_definitions',
    '公式变量定义': 'game_config.formula_variables',
    '射程配置规则': 'game_config.range_config_rules',
    '动作类型定义': 'game_config.action_type_definitions',
    '动作标记': 'game_config.action_flags',
    'Buff配置': 'game_config.buffs',
    '技能配置': 'game_config.skills',
    '技能升级消耗': 'game_config.skill_upgrade_costs',  # 改为新表名
    '动作配置': 'game_config.actions',
}

# 导入顺序（按依赖关系）
IMPORT_ORDER = [
    '角色数据类型',
    '伤害类型',
    '动作类别',
    '特征配置',
    '元效果类型定义',
    '公式变量定义',
    '射程配置规则',
    '动作类型定义',
    '动作标记',
    'Buff配置',
    '技能配置',
    '技能升级消耗',
    '动作配置',
]


class ConfigImporter:
    def __init__(self, excel_file: str, db_config: Dict):
        self.excel_file = excel_file
        self.db_config = db_config
        self.conn = None
        self.cursor = None
        self.stats = {
            'total': 0,
            'success': 0,
            'failed': 0,
            'skipped': 0,
            'details': {}
        }
    
    def connect_db(self):
        """连接数据库"""
        try:
            self.conn = psycopg2.connect(**self.db_config)
            self.cursor = self.conn.cursor()
            print("✅ 数据库连接成功")
            return True
        except Exception as e:
            print(f"❌ 数据库连接失败: {e}")
            return False
    
    def close_db(self):
        """关闭数据库连接"""
        if self.cursor:
            self.cursor.close()
        if self.conn:
            self.conn.close()
    
    def load_excel(self) -> Optional[openpyxl.Workbook]:
        """加载 Excel 文件"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            print(f"✅ Excel 文件加载成功: {self.excel_file}")
            return wb
        except Exception as e:
            print(f"❌ Excel 文件加载失败: {e}")
            return None
    
    def parse_bool(self, value: Any) -> bool:
        """解析布尔值"""
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.upper() in ['TRUE', 'YES', 'Y', '1', 'T']
        return bool(value)
    
    def parse_json(self, value: Any) -> Optional[Dict]:
        """解析 JSON 字段"""
        if not value or value == 'None':
            return None
        if isinstance(value, dict):
            return value
        if isinstance(value, str):
            try:
                return json.loads(value)
            except:
                return None
        return None
    
    def parse_tags(self, value: Any) -> Optional[List[str]]:
        """解析标签列表"""
        if not value or value == 'None':
            return None
        if isinstance(value, list):
            return value
        if isinstance(value, str):
            return [tag.strip() for tag in value.split(',') if tag.strip()]
        return None
    
    def import_hero_attribute_types(self, sheet):
        """导入角色数据类型"""
        print("\n📋 导入角色数据类型...")
        table = 'game_config.hero_attribute_type'  # 注意：表名是单数
        
        # 读取表头
        headers = [cell.value for cell in sheet[1] if cell.value]
        
        # 清空现有数据
        self.cursor.execute(f"DELETE FROM {table}")
        
        count = 0
        for row_idx in range(2, sheet.max_row + 1):
            row = sheet[row_idx]
            data = {}
            
            for idx, header in enumerate(headers):
                cell_value = row[idx].value
                
                if header == '代码':
                    data['attribute_code'] = cell_value
                elif header == '名称':
                    data['attribute_name'] = cell_value
                elif header == '分类':
                    data['category'] = cell_value
                elif header == '数据类型':
                    data['data_type'] = cell_value
                elif header == '最小值':
                    data['min_value'] = cell_value if cell_value else None
                elif header == '最大值':
                    data['max_value'] = cell_value if cell_value else None
                elif header == '默认值':
                    data['default_value'] = cell_value if cell_value else None
                elif header == '单位':
                    data['unit'] = cell_value if cell_value and cell_value != 'None' else None
                elif header == '描述':
                    data['description'] = cell_value
                elif header == '是否启用':
                    data['is_active'] = self.parse_bool(cell_value)
            
            if not data.get('attribute_code'):
                continue
            
            try:
                sql = f"""
                    INSERT INTO {table} (
                        attribute_code, attribute_name, category, data_type,
                        min_value, max_value, default_value, unit, description,
                        is_active, created_at, updated_at
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW()
                    )
                """
                self.cursor.execute(sql, (
                    data['attribute_code'], data['attribute_name'], data['category'],
                    data['data_type'], data.get('min_value'), data.get('max_value'),
                    data.get('default_value'), data.get('unit'), data.get('description'),
                    data.get('is_active', True)
                ))
                count += 1
            except Exception as e:
                print(f"  ⚠️ 行 {row_idx} 导入失败: {e}")
                continue
        
        self.conn.commit()
        print(f"  ✅ 成功导入 {count} 条记录")
        return count
    
    def import_damage_types(self, sheet):
        """导入伤害类型"""
        print("\n📋 导入伤害类型...")
        table = 'game_config.damage_types'
        
        headers = [cell.value for cell in sheet[1] if cell.value]
        self.cursor.execute(f"DELETE FROM {table}")
        
        count = 0
        for row_idx in range(2, sheet.max_row + 1):
            row = sheet[row_idx]
            data = {}
            
            for idx, header in enumerate(headers):
                cell_value = row[idx].value
                
                if header == '代码':
                    data['code'] = cell_value
                elif header == '名称':
                    data['name'] = cell_value
                elif header == '分类':
                    data['category'] = cell_value
                elif header == '抗性属性':
                    # 数据库字段名是 resistance_attribute_code
                    data['resistance_attribute_code'] = cell_value if cell_value and cell_value != 'None' else None
                elif header == '伤害减免属性':
                    # 数据库字段名是 damage_reduction_attribute_code
                    data['damage_reduction_attribute_code'] = cell_value if cell_value and cell_value != 'None' else None
                elif header == '抗性上限':
                    data['resistance_cap'] = cell_value if cell_value else None
                elif header == '颜色':
                    data['color'] = cell_value if cell_value and cell_value != 'None' else None
                elif header == '图标':
                    # 数据库字段名是 icon (不是 icon_url)
                    data['icon'] = cell_value if cell_value and cell_value != 'None' else None
                elif header == '描述':
                    data['description'] = cell_value
                elif header == '是否启用':
                    data['is_active'] = self.parse_bool(cell_value)
            
            if not data.get('code'):
                continue
            
            try:
                sql = f"""
                    INSERT INTO {table} (
                        code, name, category, resistance_attribute_code, damage_reduction_attribute_code,
                        resistance_cap, color, icon, description, is_active,
                        created_at, updated_at
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW()
                    )
                """
                self.cursor.execute(sql, (
                    data['code'], data['name'], data['category'],
                    data.get('resistance_attribute_code'), data.get('damage_reduction_attribute_code'),
                    data.get('resistance_cap'), data.get('color'), data.get('icon'),
                    data.get('description'), data.get('is_active', True)
                ))
                count += 1
            except Exception as e:
                print(f"  ⚠️  行 {row_idx} 导入失败: {e}")
                self.conn.rollback()
                continue
        
        self.conn.commit()
        print(f"  ✅ 成功导入 {count} 条记录")
        return count
    
    def import_action_categories(self, sheet):
        """导入动作类别"""
        print("\n📋 导入动作类别...")
        table = 'game_config.action_categories'
        
        headers = [cell.value for cell in sheet[1] if cell.value]
        self.cursor.execute(f"DELETE FROM {table}")
        
        count = 0
        for row_idx in range(2, sheet.max_row + 1):
            row = sheet[row_idx]
            data = {}
            
            for idx, header in enumerate(headers):
                cell_value = row[idx].value
                
                if header == '代码':
                    data['category_code'] = cell_value
                elif header == '名称':
                    data['category_name'] = cell_value
                elif header == '描述':
                    data['description'] = cell_value
                elif header == '是否启用':
                    data['is_active'] = self.parse_bool(cell_value)
            
            if not data.get('category_code'):
                continue
            
            try:
                sql = f"""
                    INSERT INTO {table} (
                        category_code, category_name, description, is_active,
                        created_at, updated_at
                    ) VALUES (%s, %s, %s, %s, NOW(), NOW())
                """
                self.cursor.execute(sql, (
                    data['category_code'], data['category_name'],
                    data.get('description'), data.get('is_active', True)
                ))
                count += 1
            except Exception as e:
                print(f"  ⚠️ 行 {row_idx} 导入失败: {e}")
                self.conn.rollback()
                continue
        
        self.conn.commit()
        print(f"  ✅ 成功导入 {count} 条记录")
        return count
    
    def import_skills(self, sheet):
        """导入技能配置"""
        print("\n📋 导入技能配置...")
        table = 'game_config.skills'
        
        headers = [cell.value for cell in sheet[1] if cell.value]
        self.cursor.execute(f"DELETE FROM {table}")
        
        count = 0
        for row_idx in range(2, sheet.max_row + 1):
            row = sheet[row_idx]
            data = {}
            
            for idx, header in enumerate(headers):
                cell_value = row[idx].value
                
                if header == '代码':
                    data['skill_code'] = cell_value
                elif header == '名称':
                    data['skill_name'] = cell_value
                elif header == '技能类型':
                    data['skill_type'] = cell_value.lower() if cell_value else 'weapon'
                elif header == '最大等级':
                    data['max_level'] = cell_value if cell_value else 10
                elif header == '特征标签':
                    # 数据库字段名是 feature_tags (text[])
                    data['feature_tags'] = self.parse_tags(cell_value)
                elif header == '被动效果':
                    data['passive_effects'] = self.parse_json(cell_value)
                elif header == '升级类型':
                    # 新增：linear/percentage/fixed
                    data['level_scaling_type'] = cell_value.lower() if cell_value else 'linear'
                elif header == '升级配置':
                    # 新增：JSON格式的升级规则
                    data['level_scaling_config'] = self.parse_json(cell_value)
                elif header == '描述':
                    data['description'] = cell_value
                elif header == '是否启用':
                    data['is_active'] = self.parse_bool(cell_value)
            
            if not data.get('skill_code'):
                continue
            
            # 设置默认值
            if 'level_scaling_type' not in data:
                data['level_scaling_type'] = 'linear'
            if 'level_scaling_config' not in data:
                data['level_scaling_config'] = {}
            
            try:
                sql = f"""
                    INSERT INTO {table} (
                        skill_code, skill_name, skill_type, max_level, feature_tags,
                        passive_effects, level_scaling_type, level_scaling_config,
                        description, is_active, created_at, updated_at
                    ) VALUES (%s, %s, %s::skill_type_enum, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                """
                self.cursor.execute(sql, (
                    data['skill_code'], data['skill_name'], data.get('skill_type', 'weapon'),
                    data.get('max_level', 10), data.get('feature_tags'),
                    json.dumps(data.get('passive_effects')) if data.get('passive_effects') else None,
                    data.get('level_scaling_type', 'linear'),
                    json.dumps(data.get('level_scaling_config', {})),
                    data.get('description'), data.get('is_active', True)
                ))
                count += 1
            except Exception as e:
                print(f"  ⚠️ 行 {row_idx} 导入失败: {e}")
                self.conn.rollback()
                continue
        
        self.conn.commit()
        print(f"  ✅ 成功导入 {count} 条记录")
        return count
    
    def import_skill_upgrade_costs(self, sheet):
        """导入技能升级消耗配置（全局）"""
        print("\n📋 导入技能升级消耗配置...")
        table = 'game_config.skill_upgrade_costs'
        
        headers = [cell.value for cell in sheet[1] if cell.value]
        self.cursor.execute(f"DELETE FROM {table}")
        
        count = 0
        for row_idx in range(2, sheet.max_row + 1):
            row = sheet[row_idx]
            data = {}
            
            for idx, header in enumerate(headers):
                cell_value = row[idx].value
                
                if header == '等级' or header == '升级到等级':
                    data['level_number'] = cell_value
                elif header == '经验消耗' or header == 'XP消耗':
                    data['cost_xp'] = cell_value if cell_value else 0
                elif header == '金币消耗':
                    data['cost_gold'] = cell_value if cell_value else 0
                elif header == '材料消耗':
                    # JSON格式：[{"item_code": "xxx", "count": 5}]
                    data['cost_materials'] = self.parse_json(cell_value) if cell_value else []
            
            if not data.get('level_number'):
                continue
            
            try:
                sql = f"""
                    INSERT INTO {table} (
                        level_number, cost_xp, cost_gold, cost_materials,
                        created_at, updated_at
                    ) VALUES (%s, %s, %s, %s, NOW(), NOW())
                """
                self.cursor.execute(sql, (
                    data['level_number'], 
                    data.get('cost_xp', 0),
                    data.get('cost_gold', 0),
                    json.dumps(data.get('cost_materials', []))
                ))
                count += 1
            except Exception as e:
                print(f"  ⚠️ 行 {row_idx} 导入失败: {e}")
                self.conn.rollback()
                continue
        
        self.conn.commit()
        print(f"  ✅ 成功导入 {count} 条记录")
        return count
    
    def import_actions(self, sheet):
        """导入动作配置"""
        print("\n📋 导入动作配置...")
        table = 'game_config.actions'
        
        headers = [cell.value for cell in sheet[1] if cell.value]
        self.cursor.execute(f"DELETE FROM {table}")
        
        count = 0
        for row_idx in range(2, sheet.max_row + 1):
            row = sheet[row_idx]
            data = {}
            
            for idx, header in enumerate(headers):
                cell_value = row[idx].value
                
                if header == '动作代码':
                    data['action_code'] = cell_value
                elif header == '动作名称':
                    data['action_name'] = cell_value
                elif header == '动作类别':
                    # 这是 action_categories 的代码,暂不处理外键
                    data['action_category_code'] = cell_value
                elif header == '动作类型':
                    data['action_type'] = cell_value.lower() if cell_value else 'main'
                elif header == '特征标签':
                    # 数据库字段名是 feature_tags (text[])
                    data['feature_tags'] = self.parse_tags(cell_value)
                elif header == '关联技能':
                    # 这是技能代码,暂不处理外键
                    data['related_skill_code'] = cell_value if cell_value and cell_value != 'None' else None
                elif header == '射程配置':
                    # 射程配置为JSON
                    data['range_config'] = self.parse_json(cell_value) if cell_value else {}
                elif header == '开始标记':
                    # 分号分隔转为列表
                    if cell_value and cell_value != 'None':
                        data['start_flags'] = [f.strip() for f in cell_value.split(';') if f.strip()]
                    else:
                        data['start_flags'] = None
                elif header == '需求条件':
                    data['requirements'] = self.parse_json(cell_value)
                elif header == '描述':
                    data['description'] = cell_value
            
            if not data.get('action_code'):
                continue
            
            try:
                sql = f"""
                    INSERT INTO {table} (
                        action_code, action_name, action_type, feature_tags,
                        range_config, start_flags, requirements, description, 
                        is_active, created_at, updated_at
                    ) VALUES (%s, %s, %s::action_type_enum, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                """
                self.cursor.execute(sql, (
                    data['action_code'], data['action_name'], data.get('action_type', 'main'),
                    data.get('feature_tags'),
                    json.dumps(data.get('range_config', {})),
                    data.get('start_flags'),
                    json.dumps(data.get('requirements')) if data.get('requirements') else None,
                    data.get('description'), True
                ))
                count += 1
            except Exception as e:
                print(f"  ⚠️ 行 {row_idx} 导入失败: {e}")
                self.conn.rollback()
                continue
        
        self.conn.commit()
        print(f"  ✅ 成功导入 {count} 条记录")
        return count
    
    def import_buffs(self, sheet):
        """导入Buff配置"""
        print("\n📋 导入Buff配置...")
        table = 'game_config.buffs'
        
        headers = [cell.value for cell in sheet[1] if cell.value]
        self.cursor.execute(f"DELETE FROM {table}")
        
        count = 0
        for row_idx in range(2, sheet.max_row + 1):
            row = sheet[row_idx]
            data = {}
            
            for idx, header in enumerate(headers):
                cell_value = row[idx].value
                
                if header == '代码':
                    data['buff_code'] = cell_value
                elif header == '名称':
                    data['buff_name'] = cell_value
                elif header == 'Buff类型':
                    data['buff_type'] = cell_value if cell_value else 'buff'
                elif header == '分类':
                    data['category'] = cell_value if cell_value and cell_value != 'None' else None
                elif header == '特征标签':
                    # 数据库字段名是 feature_tags (text[])
                    data['feature_tags'] = self.parse_tags(cell_value)
                elif header == '默认持续时间':
                    data['default_duration'] = cell_value if cell_value else 1
                elif header == '效果描述':
                    data['effect_description'] = cell_value
                elif header == '叠加规则':
                    data['stack_rule'] = cell_value if cell_value else 'no_stack'
                elif header == '最大叠加层数':
                    data['max_stacks'] = cell_value if cell_value else 1
                elif header == '描述':
                    data['description'] = cell_value
                elif header == '是否启用':
                    data['is_active'] = self.parse_bool(cell_value)
            
            if not data.get('buff_code'):
                continue
            
            try:
                sql = f"""
                    INSERT INTO {table} (
                        buff_code, buff_name, buff_type, category, feature_tags,
                        default_duration, effect_description, stack_rule, max_stacks,
                        description, is_active, created_at, updated_at
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                """
                self.cursor.execute(sql, (
                    data['buff_code'], data['buff_name'], data.get('buff_type', 'buff'),
                    data.get('category'), data.get('feature_tags'), data.get('default_duration', 1),
                    data.get('effect_description'), data.get('stack_rule', 'no_stack'),
                    data.get('max_stacks', 1), data.get('description'),
                    data.get('is_active', True)
                ))
                count += 1
            except Exception as e:
                print(f"  ⚠️ 行 {row_idx} 导入失败: {e}")
                self.conn.rollback()
                continue
        
        self.conn.commit()
        print(f"  ✅ 成功导入 {count} 条记录")
        return count
    
    def import_effect_type_definitions(self, sheet):
        """导入元效果类型定义"""
        print("\n📋 导入元效果类型定义...")
        table = 'game_config.effect_type_definitions'
        
        headers = [cell.value for cell in sheet[1] if cell.value]
        self.cursor.execute(f"DELETE FROM {table}")
        
        count = 0
        for row_idx in range(2, sheet.max_row + 1):
            row = sheet[row_idx]
            data = {}
            
            for idx, header in enumerate(headers):
                cell_value = row[idx].value
                
                if header == '效果类型代码':
                    data['effect_type_code'] = cell_value
                elif header == '效果类型名称':
                    data['effect_type_name'] = cell_value
                elif header == '描述':
                    data['description'] = cell_value
                elif header == '参数列表':
                    # text[] 数组字段
                    data['parameter_list'] = self.parse_tags(cell_value)
                elif header == '参数说明':
                    data['parameter_descriptions'] = cell_value
                elif header == '参数定义':
                    # JSONB字段
                    data['parameter_definitions'] = self.parse_json(cell_value)
                elif header == '失败处理':
                    data['failure_handling'] = cell_value if cell_value and cell_value != 'None' else None
                elif header == 'JSON模板':
                    # JSONB字段
                    data['json_template'] = self.parse_json(cell_value)
                elif header == '示例':
                    data['example'] = cell_value
                elif header == '备注':
                    data['notes'] = cell_value
                elif header == '是否启用':
                    data['is_active'] = self.parse_bool(cell_value)
            
            if not data.get('effect_type_code'):
                continue
            
            try:
                sql = f"""
                    INSERT INTO {table} (
                        effect_type_code, effect_type_name, description, parameter_list,
                        parameter_descriptions, parameter_definitions, failure_handling,
                        json_template, example, notes, is_active, created_at, updated_at
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                """
                self.cursor.execute(sql, (
                    data['effect_type_code'], data['effect_type_name'], data.get('description'),
                    data.get('parameter_list'), data.get('parameter_descriptions'),
                    json.dumps(data.get('parameter_definitions')) if data.get('parameter_definitions') else None,
                    data.get('failure_handling'),
                    json.dumps(data.get('json_template')) if data.get('json_template') else None,
                    data.get('example'), data.get('notes'), data.get('is_active', True)
                ))
                count += 1
            except Exception as e:
                print(f"  ⚠️ 行 {row_idx} 导入失败: {e}")
                self.conn.rollback()
                continue
        
        self.conn.commit()
        print(f"  ✅ 成功导入 {count} 条记录")
        return count
    
    def import_formula_variables(self, sheet):
        """导入公式变量定义"""
        print("\n📋 导入公式变量定义...")
        table = 'game_config.formula_variables'
        
        headers = [cell.value for cell in sheet[1] if cell.value]
        self.cursor.execute(f"DELETE FROM {table}")
        
        count = 0
        for row_idx in range(2, sheet.max_row + 1):
            row = sheet[row_idx]
            data = {}
            
            for idx, header in enumerate(headers):
                cell_value = row[idx].value
                
                if header == '变量代码':
                    data['variable_code'] = cell_value
                elif header == '变量名称':
                    data['variable_name'] = cell_value
                elif header == '变量类型':
                    data['variable_type'] = cell_value
                elif header == '作用域':
                    data['scope'] = cell_value
                elif header == '数据类型':
                    data['data_type'] = cell_value
                elif header == '描述':
                    data['description'] = cell_value
                elif header == '示例':
                    data['example'] = cell_value
                elif header == '是否启用':
                    data['is_active'] = self.parse_bool(cell_value)
            
            if not data.get('variable_code'):
                continue
            
            try:
                sql = f"""
                    INSERT INTO {table} (
                        variable_code, variable_name, variable_type, scope, data_type,
                        description, example, is_active, created_at, updated_at
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                """
                self.cursor.execute(sql, (
                    data['variable_code'], data['variable_name'], data['variable_type'],
                    data['scope'], data['data_type'], data.get('description'),
                    data.get('example'), data.get('is_active', True)
                ))
                count += 1
            except Exception as e:
                print(f"  ⚠️ 行 {row_idx} 导入失败: {e}")
                self.conn.rollback()
                continue
        
        self.conn.commit()
        print(f"  ✅ 成功导入 {count} 条记录")
        return count
    
    def import_range_config_rules(self, sheet):
        """导入射程配置规则"""
        print("\n📋 导入射程配置规则...")
        table = 'game_config.range_config_rules'
        
        headers = [cell.value for cell in sheet[1] if cell.value]
        self.cursor.execute(f"DELETE FROM {table}")
        
        count = 0
        for row_idx in range(2, sheet.max_row + 1):
            row = sheet[row_idx]
            data = {}
            
            for idx, header in enumerate(headers):
                cell_value = row[idx].value
                
                if header == '参数类型':
                    data['parameter_type'] = cell_value
                elif header == '参数格式':
                    data['parameter_format'] = cell_value
                elif header == '描述':
                    data['description'] = cell_value
                elif header == '示例':
                    data['example'] = cell_value
                elif header == '备注':
                    data['notes'] = cell_value
                elif header == '是否启用':
                    data['is_active'] = self.parse_bool(cell_value)
            
            if not data.get('parameter_type'):
                continue
            
            try:
                sql = f"""
                    INSERT INTO {table} (
                        parameter_type, parameter_format, description, example, notes,
                        is_active, created_at
                    ) VALUES (%s, %s, %s, %s, %s, %s, NOW())
                """
                self.cursor.execute(sql, (
                    data['parameter_type'], data['parameter_format'], data.get('description'),
                    data.get('example'), data.get('notes'), data.get('is_active', True)
                ))
                count += 1
            except Exception as e:
                print(f"  ⚠️ 行 {row_idx} 导入失败: {e}")
                self.conn.rollback()
                continue
        
        self.conn.commit()
        print(f"  ✅ 成功导入 {count} 条记录")
        return count
    
    def import_action_type_definitions(self, sheet):
        """导入动作类型定义"""
        print("\n📋 导入动作类型定义...")
        table = 'game_config.action_type_definitions'
        
        headers = [cell.value for cell in sheet[1] if cell.value]
        self.cursor.execute(f"DELETE FROM {table}")
        
        count = 0
        for row_idx in range(2, sheet.max_row + 1):
            row = sheet[row_idx]
            data = {}
            
            for idx, header in enumerate(headers):
                cell_value = row[idx].value
                
                if header == '动作类型':
                    data['action_type'] = cell_value
                elif header == '描述':
                    data['description'] = cell_value
                elif header == '每回合限制':
                    data['per_turn_limit'] = cell_value if cell_value else None
                elif header == '使用时机':
                    data['usage_timing'] = cell_value
                elif header == '示例':
                    data['example'] = cell_value
                elif header == '是否启用':
                    data['is_active'] = self.parse_bool(cell_value)
            
            if not data.get('action_type'):
                continue
            
            try:
                sql = f"""
                    INSERT INTO {table} (
                        action_type, description, per_turn_limit, usage_timing, example,
                        is_active, created_at
                    ) VALUES (%s, %s, %s, %s, %s, %s, NOW())
                """
                self.cursor.execute(sql, (
                    data['action_type'], data.get('description'), data.get('per_turn_limit'),
                    data.get('usage_timing'), data.get('example'), data.get('is_active', True)
                ))
                count += 1
            except Exception as e:
                print(f"  ⚠️ 行 {row_idx} 导入失败: {e}")
                self.conn.rollback()
                continue
        
        self.conn.commit()
        print(f"  ✅ 成功导入 {count} 条记录")
        return count
    
    def import_action_flags(self, sheet):
        """导入动作标记"""
        print("\n📋 导入动作标记...")
        table = 'game_config.action_flags'
        
        headers = [cell.value for cell in sheet[1] if cell.value]
        self.cursor.execute(f"DELETE FROM {table}")
        
        count = 0
        for row_idx in range(2, sheet.max_row + 1):
            row = sheet[row_idx]
            data = {}
            
            for idx, header in enumerate(headers):
                cell_value = row[idx].value
                
                if header == '标记代码':
                    data['flag_code'] = cell_value
                elif header == '标记名称':
                    data['flag_name'] = cell_value
                elif header == '分类':
                    data['category'] = cell_value if cell_value and cell_value != 'None' else None
                elif header == '持续类型':
                    data['duration_type'] = cell_value if cell_value else 'action'
                elif header == '默认持续时间':
                    data['default_duration'] = str(cell_value) if cell_value else '1'
                elif header == '自动移除条件':
                    data['auto_remove_condition'] = cell_value if cell_value and cell_value != 'None' else None
                elif header == '移除事件':
                    # text[] 数组
                    data['remove_on_events'] = self.parse_tags(cell_value)
                elif header == '是否可见':
                    data['is_visible'] = self.parse_bool(cell_value)
                elif header == '是否可叠加':
                    data['is_stackable'] = self.parse_bool(cell_value)
                elif header == '最大叠加层数':
                    data['max_stacks'] = cell_value if cell_value else 1
                elif header == '提供优势':
                    data['provides_advantage'] = self.parse_bool(cell_value)
                elif header == '提供劣势':
                    data['provides_disadvantage'] = self.parse_bool(cell_value)
                elif header == '优势适用于':
                    # text[] 数组
                    data['advantage_applies_to'] = self.parse_tags(cell_value)
                elif header == '标记效果':
                    # JSONB
                    data['flag_effects'] = self.parse_json(cell_value)
                elif header == '修正效果':
                    # JSONB
                    data['modifier_effects'] = self.parse_json(cell_value)
                elif header == '描述':
                    data['description'] = cell_value
                elif header == '是否启用':
                    data['is_active'] = self.parse_bool(cell_value)
            
            if not data.get('flag_code'):
                continue
            
            try:
                sql = f"""
                    INSERT INTO {table} (
                        flag_code, flag_name, category, duration_type, default_duration,
                        auto_remove_condition, remove_on_events, is_visible, is_stackable,
                        max_stacks, provides_advantage, provides_disadvantage,
                        advantage_applies_to, flag_effects, modifier_effects, description,
                        is_active, created_at, updated_at
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                """
                self.cursor.execute(sql, (
                    data['flag_code'], data['flag_name'], data.get('category'),
                    data.get('duration_type', 'action'), data.get('default_duration', '1'),
                    data.get('auto_remove_condition'), data.get('remove_on_events'),
                    data.get('is_visible', False), data.get('is_stackable', False),
                    data.get('max_stacks', 1), data.get('provides_advantage', False),
                    data.get('provides_disadvantage', False), data.get('advantage_applies_to'),
                    json.dumps(data.get('flag_effects')) if data.get('flag_effects') else None,
                    json.dumps(data.get('modifier_effects')) if data.get('modifier_effects') else None,
                    data.get('description'), data.get('is_active', True)
                ))
                count += 1
            except Exception as e:
                print(f"  ⚠️ 行 {row_idx} 导入失败: {e}")
                self.conn.rollback()
                continue
        
        self.conn.commit()
        print(f"  ✅ 成功导入 {count} 条记录")
        return count
    
    def import_tags(self, sheet):
        """导入特征配置"""
        print("\n📋 导入特征配置...")
        table = 'game_config.tags'
        
        headers = [cell.value for cell in sheet[1] if cell.value]
        self.cursor.execute(f"DELETE FROM {table}")
        
        # 定义分类映射：适用类型 -> category 枚举值
        # 实际数据库枚举值：class, item, skill, monster
        category_map = {
            '技能': 'skill',
            'skill': 'skill',
            '物品': 'item',
            'item': 'item',
            '职业': 'class',
            'class': 'class',
            '怪物': 'monster',
            'monster': 'monster'
        }
        
        count = 0
        for row_idx in range(2, sheet.max_row + 1):
            row = sheet[row_idx]
            data = {}
            
            for idx, header in enumerate(headers):
                cell_value = row[idx].value
                
                if header == '特征代码':
                    data['tag_code'] = cell_value
                elif header == '特征名称':
                    data['tag_name'] = cell_value
                elif header == '适用类型':
                    # 数据库使用 category 字段（枚举类型：class, item, skill, monster）
                    types_str = str(cell_value) if cell_value else ''
                    # 默认为 skill
                    data['category'] = 'skill'
                    # 尝试从字符串中映射
                    for key, value in category_map.items():
                        if key in types_str:
                            data['category'] = value
                            break
                elif header == '颜色':
                    data['color'] = cell_value if cell_value and cell_value != 'None' else None
                elif header == '图标':
                    # 数据库字段名是 icon (不是 icon_url)
                    data['icon'] = cell_value if cell_value and cell_value != 'None' else None
                elif header == '描述':
                    data['description'] = cell_value
                elif header == '是否启用':
                    data['is_active'] = self.parse_bool(cell_value)
                elif header == '显示顺序':
                    data['display_order'] = cell_value if cell_value else 0
            
            if not data.get('tag_code'):
                continue
            
            # 设置默认值
            if 'display_order' not in data:
                data['display_order'] = 0
            if 'category' not in data:
                data['category'] = 'skill'  # 默认为 skill
            
            try:
                sql = f"""
                    INSERT INTO {table} (
                        tag_code, tag_name, category, color, icon,
                        description, is_active, display_order, created_at, updated_at
                    ) VALUES (%s, %s, %s::tag_type_enum, %s, %s, %s, %s, %s, NOW(), NOW())
                """
                self.cursor.execute(sql, (
                    data['tag_code'], data['tag_name'], data['category'],
                    data.get('color'), data.get('icon'), data.get('description'),
                    data.get('is_active', True), data.get('display_order', 0)
                ))
                count += 1
            except Exception as e:
                print(f"  ⚠️ 行 {row_idx} 导入失败: {e}")
                self.conn.rollback()
                continue
        
        self.conn.commit()
        print(f"  ✅ 成功导入 {count} 条记录")
        return count
    
    def run(self):
        """执行导入"""
        print("="*80)
        print(" 🎮 游戏配置导入工具")
        print("="*80)
        
        # 加载 Excel
        wb = self.load_excel()
        if not wb:
            return False
        
        # 连接数据库
        if not self.connect_db():
            return False
        
        try:
            # 按顺序导入
            for sheet_name in IMPORT_ORDER:
                if sheet_name not in wb.sheetnames:
                    print(f"\n⚠️ Sheet '{sheet_name}' 不存在，跳过")
                    continue
                
                sheet = wb[sheet_name]
                count = 0
                
                # 根据 Sheet 名称调用对应的导入方法
                try:
                    if sheet_name == '角色数据类型':
                        count = self.import_hero_attribute_types(sheet)
                    elif sheet_name == '伤害类型':
                        count = self.import_damage_types(sheet)
                    elif sheet_name == '动作类别':
                        count = self.import_action_categories(sheet)
                    elif sheet_name == '特征配置':
                        count = self.import_tags(sheet)
                    elif sheet_name == '元效果类型定义':
                        count = self.import_effect_type_definitions(sheet)
                    elif sheet_name == '公式变量定义':
                        count = self.import_formula_variables(sheet)
                    elif sheet_name == '射程配置规则':
                        count = self.import_range_config_rules(sheet)
                    elif sheet_name == '动作类型定义':
                        count = self.import_action_type_definitions(sheet)
                    elif sheet_name == '动作标记':
                        count = self.import_action_flags(sheet)
                    elif sheet_name == 'Buff配置':
                        count = self.import_buffs(sheet)
                    elif sheet_name == '技能配置':
                        count = self.import_skills(sheet)
                    elif sheet_name == '技能升级消耗':
                        count = self.import_skill_upgrade_costs(sheet)
                    elif sheet_name == '动作配置':
                        count = self.import_actions(sheet)
                    else:
                        print(f"\n⚠️ Sheet '{sheet_name}' 暂未实现导入逻辑，跳过")
                        count = 0
                except Exception as e:
                    print(f"\n❌ 导入 '{sheet_name}' 失败: {e}")
                    count = 0
                
                self.stats['total'] += count
            
            print("\n" + "="*80)
            print("📊 导入统计:")
            print(f"  总计导入: {self.stats['total']} 条记录")
            print("="*80)
            
            return True
        
        except Exception as e:
            print(f"\n❌ 导入过程发生错误: {e}")
            self.conn.rollback()
            return False
        finally:
            self.close_db()


def main():
    parser = argparse.ArgumentParser(description='游戏配置表导入工具')
    parser.add_argument('--file', default='configs/game/游戏配置表_v2.0.0.xlsx',
                        help='Excel 配置文件路径')
    parser.add_argument('--host', default='localhost', help='数据库主机')
    parser.add_argument('--port', default=5432, type=int, help='数据库端口')
    parser.add_argument('--user', default='tsu_user', help='数据库用户')
    parser.add_argument('--password', default='tsu_password', help='数据库密码')
    parser.add_argument('--database', default='tsu_db', help='数据库名')
    
    args = parser.parse_args()
    
    # 更新数据库配置
    db_config = {
        'host': args.host,
        'port': args.port,
        'database': args.database,
        'user': args.user,
        'password': args.password
    }
    
    # 创建导入器并执行
    importer = ConfigImporter(args.file, db_config)
    success = importer.run()
    
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
